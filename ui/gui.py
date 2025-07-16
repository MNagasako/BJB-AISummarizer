import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# DummyApp: GUIエントリーポイント
class DummyApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BJB-AISummarizer ダミーUI")
        self.geometry("800x600")

        tk.Label(self, text="XLSXファイルを選択してください").pack(pady=10)
        tk.Button(self, text="ファイル選択", command=self.select_file).pack(pady=5)

        self.col_label = tk.Label(self, text="カラム一覧: -")
        self.col_label.pack(pady=5)

        self.tree = ttk.Treeview(self, show="headings")
        self.tree.pack(expand=True, fill="both", padx=10, pady=10)

        tk.Button(self, text="出力列リスト編集", command=self.edit_output_columns).pack(pady=5)
        tk.Button(self, text="実行", command=self.run_process).pack(pady=5)
        tk.Button(self, text="終了", command=self.quit).pack(pady=5)

        self.input_columns = []
        self.output_columns = []  # [{name, type, source}]
        self.input_rows = []

    def select_file(self):
        from openpyxl import load_workbook
        file_path = tk.filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    raise Exception("空のファイルです")
                self.input_columns = list(rows[0])
                self.input_rows = rows[1:]
                self.col_label.config(text=f"カラム一覧: {', '.join(str(c) for c in self.input_columns)}")
                self.show_rows(self.input_columns, self.input_rows[:20])
            except Exception as e:
                messagebox.showerror("エラー", f"ファイル読み込み失敗: {e}")

    def show_rows(self, columns, data_rows):
        self.tree.delete(*self.tree.get_children())
        self.tree['columns'] = columns
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, anchor='center')
        for row in data_rows:
            self.tree.insert('', 'end', values=row)

    def edit_output_columns(self):
        if hasattr(self, '_output_dialog') and self._output_dialog is not None and self._output_dialog.winfo_exists():
            self._output_dialog.lift()
            self._output_dialog.focus_force()
            return
        if not self.input_columns:
            messagebox.showwarning("警告", "先に入力ファイルを選択してください")
            return
        # 初回のみ初期化
        if not hasattr(self, '_output_dialog_initialized'):
            dialog = OutputColumnDialog(self, self.input_columns, [])
            self._output_dialog_initialized = True
        else:
            dialog = OutputColumnDialog(self, self.input_columns, self.output_columns)
        self._output_dialog = dialog
        def on_close():
            self._output_dialog = None
            dialog.destroy()
        dialog.protocol("WM_DELETE_WINDOW", on_close)
        self.wait_window(dialog)
        self.output_columns = dialog.get_output_columns()

    def run_process(self):
        if not self.input_columns or not self.input_rows:
            messagebox.showwarning("警告", "入力ファイルを選択してください")
            return
        if not self.output_columns:
            messagebox.showwarning("警告", "出力列リストを編集してください")
            return
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append([col['name'] for col in self.output_columns])
        for row in self.input_rows:
            out_row = []
            row_dict = dict(zip(self.input_columns, row))
            for col in self.output_columns:
                val = row_dict.get(col['source'], "")
                if col['type'] == 'copy':
                    out_row.append(val)
                elif col['type'] == 'halfwidth':
                    out_row.append(str(val).translate(str.maketrans({chr(0xFF01+i):chr(0x21+i) for i in range(94)})))
                elif col['type'] == 'length':
                    out_row.append(len(str(val)))
                else:
                    out_row.append(val)
            ws.append(out_row)
        import os
        out_path = os.path.join(os.path.dirname(__file__), '..', 'output', 'output.xlsx')
        wb.save(out_path)
        messagebox.showinfo("完了", f"出力ファイルを保存しました: {out_path}")
# 必要なimportをファイル先頭で一度だけ
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
# 単一出力列追加ダイアログ
class SingleColumnDialog(tk.Toplevel):
    def __init__(self, parent, input_columns, col=None):
        super().__init__(parent)
        self.title("出力列追加/編集")
        self.geometry("400x260")
        self.result = None

        tk.Label(self, text="出力列名").pack(pady=5)
        self.name_entry = tk.Entry(self)
        self.name_entry.pack(pady=5)

        tk.Label(self, text="処理タイプ").pack(pady=5)
        self.type_var = tk.StringVar(value='copy')
        type_menu = ttk.Combobox(self, textvariable=self.type_var, values=['copy','halfwidth','length'], state='readonly')
        type_menu.pack(pady=5)

        tk.Label(self, text="元列").pack(pady=5)
        self.source_var = tk.StringVar(value=input_columns[0] if input_columns else '')
        source_menu = ttk.Combobox(self, textvariable=self.source_var, values=input_columns, state='readonly')
        source_menu.pack(pady=5)

        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="決定", command=self.on_ok).pack(side='left', padx=5)
        # 編集時は初期値セット
        if col:
            self.name_entry.insert(0, col.get('name', ''))
            self.type_var.set(col.get('type', 'copy'))
            self.source_var.set(col.get('source', input_columns[0] if input_columns else ''))

    def on_ok(self):
        name = self.name_entry.get()
        type_ = self.type_var.get()
        source = self.source_var.get()
        if name and type_ and source:
            self.result = {'name': name, 'type': type_, 'source': source}
            self.destroy()

    def get_column(self):
        return self.result

    def show_rows(self, columns, data_rows):
        self.tree.delete(*self.tree.get_children())
        self.tree['columns'] = columns
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, anchor='center')
        for row in data_rows:
            self.tree.insert('', 'end', values=row)

    def edit_output_columns(self):
        if not self.input_columns:
            messagebox.showwarning("警告", "先に入力ファイルを選択してください")
            return
        dialog = OutputColumnDialog(self, self.input_columns, self.output_columns)
        self.wait_window(dialog)
        self.output_columns = dialog.get_output_columns()

    def run_process(self):
        if not self.input_columns or not self.input_rows:
            messagebox.showwarning("警告", "入力ファイルを選択してください")
            return
        if not self.output_columns:
            messagebox.showwarning("警告", "出力列リストを編集してください")
            return
        # 処理ロジック（ダミー: copy/halfwidth/lengthのみ）
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # ヘッダ
        ws.append([col['name'] for col in self.output_columns])
        # データ
        for row in self.input_rows:
            out_row = []
            row_dict = dict(zip(self.input_columns, row))
            for col in self.output_columns:
                val = row_dict.get(col['source'], "")
                if col['type'] == 'copy':
                    out_row.append(val)
                elif col['type'] == 'halfwidth':
                    out_row.append(str(val).translate(str.maketrans({chr(0xFF01+i):chr(0x21+i) for i in range(94)})))
                elif col['type'] == 'length':
                    out_row.append(len(str(val)))
                else:
                    out_row.append(val)
            ws.append(out_row)
        # 保存
        import os
        out_path = os.path.join(os.path.dirname(__file__), '..', 'output', 'output.xlsx')
        wb.save(out_path)
        messagebox.showinfo("完了", f"出力ファイルを保存しました: {out_path}")

# 出力列編集ダイアログ
import json
import os
import tkinter as tk
from tkinter import ttk, messagebox

class OutputColumnDialog(tk.Toplevel):
    def __init__(self, parent, input_columns, output_columns):
        super().__init__(parent)
        self.title("出力列リスト編集")
        self.geometry("520x420")
        self.input_columns = input_columns
        self.original_output_columns = output_columns.copy()
        self.output_columns = output_columns.copy() if output_columns else [
            {'name': col, 'type': 'copy', 'source': col} for col in input_columns
        ]

        self.listbox = tk.Listbox(self)
        self.listbox.pack(fill='both', expand=True, padx=10, pady=10)
        self.refresh_list()

        self.listbox.bind('<Button-3>', self.show_context_menu)
        self.listbox.bind('<Delete>', lambda e: self.delete_column())
        self.listbox.bind('<Double-Button-1>', lambda e: self.edit_column())
        self.listbox.bind('<B1-Motion>', self.on_drag)
class OutputColumnDialog(tk.Toplevel):
    def __init__(self, parent, input_columns, output_columns):
        super().__init__(parent)
        self.title("出力列リスト編集")
        self.geometry("540x440")
        self.input_columns = input_columns
        self.original_output_columns = output_columns.copy()
        self.output_columns = output_columns.copy() if output_columns else [
            {'name': col, 'type': 'copy', 'source': col} for col in input_columns
        ]

        # Treeviewで列分け表示
        self.tree = ttk.Treeview(self, columns=('name','type','source'), show='headings', selectmode='browse')
        self.tree.heading('name', text='出力列名')
        self.tree.heading('type', text='処理タイプ')
        self.tree.heading('source', text='元列')
        self.tree.column('name', width=150)
        self.tree.column('type', width=100)
        self.tree.column('source', width=150)
        self.tree.pack(fill='both', expand=True, padx=10, pady=10)
        self.refresh_list()

        self.tree.bind('<Button-3>', self.show_context_menu)
        self.tree.bind('<Delete>', lambda e: self.delete_column())
        self.tree.bind('<Double-Button-1>', lambda e: self.edit_column())
        self.tree.bind('<B1-Motion>', self.on_drag)
        self.tree.bind('<ButtonRelease-1>', self.on_drop)
        self.drag_index = None

        frm = tk.Frame(self)
        frm.pack(pady=5)
        tk.Button(frm, text="追加", command=self.add_column).pack(side='left', padx=5)
        tk.Button(frm, text="削除", command=self.delete_column).pack(side='left', padx=5)
        tk.Button(frm, text="編集", command=self.edit_column).pack(side='left', padx=5)
        tk.Button(frm, text="全コピー", command=self.copy_all).pack(side='left', padx=5)
        tk.Button(frm, text="一括クリア", command=self.clear_all).pack(side='left', padx=5)
        tk.Button(frm, text="復元", command=self.restore_all).pack(side='left', padx=5)
        tk.Button(frm, text="保存", command=self.save_setting).pack(side='left', padx=5)
        tk.Button(frm, text="読込", command=self.load_setting).pack(side='left', padx=5)
        tk.Button(frm, text="閉じる", command=self.destroy).pack(side='left', padx=5)

        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="編集", command=self.edit_column)
        self.context_menu.add_command(label="削除", command=self.delete_column)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="保存", command=self.save_setting)
        self.context_menu.add_command(label="読込", command=self.load_setting)

    def refresh_list(self):
        self.tree.delete(*self.tree.get_children())
        for col in self.output_columns:
            self.tree.insert('', 'end', values=(col['name'], col['type'], col['source']))

    def add_column(self):
        dialog = SingleColumnDialog(self, self.input_columns)
        self.wait_window(dialog)
        col = dialog.get_column()
        if col:
            self.output_columns.append(col)
            self.refresh_list()

    def delete_column(self):
        sel = self.tree.selection()
        if sel:
            idx = self.tree.index(sel[0])
            del self.output_columns[idx]
            self.refresh_list()

    def edit_column(self):
        sel = self.tree.selection()
        if sel:
            idx = self.tree.index(sel[0])
            col = self.output_columns[idx]
            dialog = SingleColumnDialog(self, self.input_columns, col)
            self.wait_window(dialog)
            new_col = dialog.get_column()
            if new_col:
                self.output_columns[idx] = new_col
                self.refresh_list()

    def copy_all(self):
        self.output_columns = [
            {'name': col, 'type': 'copy', 'source': col} for col in self.input_columns
        ]
        self.refresh_list()

    def clear_all(self):
        self.output_columns = []
        self.refresh_list()

    def restore_all(self):
        self.output_columns = self.original_output_columns.copy()
        self.refresh_list()

    def save_setting(self):
        path = filedialog.asksaveasfilename(
            initialdir=os.path.join(os.path.dirname(__file__), '..', 'config'),
            defaultextension='.json',
            filetypes=[('JSON files', '*.json')],
            title='出力列設定の保存')
        if not path:
            return
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(self.output_columns, f, ensure_ascii=False, indent=2)
            messagebox.showinfo('保存', f'設定を保存しました: {path}')
        except Exception as e:
            messagebox.showerror('保存失敗', str(e))

    def load_setting(self):
        path = filedialog.askopenfilename(
            initialdir=os.path.join(os.path.dirname(__file__), '..', 'config'),
            defaultextension='.json',
            filetypes=[('JSON files', '*.json')],
            title='出力列設定の読込')
        if not path:
            return
        try:
            with open(path, 'r', encoding='utf-8') as f:
                loaded = json.load(f)
            # 元データにない列は警告のみ
            missing = [col for col in loaded if col['source'] not in self.input_columns]
            if missing:
                messagebox.showwarning('警告', f'元データにない列があります: {[col['source'] for col in missing]}')
            self.output_columns = loaded
            self.refresh_list()
            messagebox.showinfo('読込', f'設定を読込ました: {path}')
        except Exception as e:
            messagebox.showerror('読込失敗', str(e))

    def show_context_menu(self, event):
        try:
            self.tree.selection_remove(self.tree.selection())
            item = self.tree.identify_row(event.y)
            if item:
                self.tree.selection_set(item)
        except:
            pass
        self.context_menu.tk_popup(event.x_root, event.y_root)

    def on_drag(self, event):
        item = self.tree.identify_row(event.y)
        if item and self.drag_index is None:
            self.drag_index = self.tree.index(item)

    def on_drop(self, event):
        if self.drag_index is not None:
            item = self.tree.identify_row(event.y)
            if item:
                drop_idx = self.tree.index(item)
                if drop_idx != self.drag_index:
                    col = self.output_columns.pop(self.drag_index)
                    self.output_columns.insert(drop_idx, col)
                    self.refresh_list()
            self.drag_index = None

    def get_output_columns(self):
        return self.output_columns
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="追加/決定", command=self.on_ok).pack(side='left', padx=5)

        # 編集時は初期値セット
        if col:
            self.name_entry.insert(0, col.get('name', ''))
            self.type_var.set(col.get('type', 'copy'))
            self.source_var.set(col.get('source', input_columns[0] if input_columns else ''))

    def on_ok(self):
        name = self.name_entry.get()
        type_ = self.type_var.get()
        source = self.source_var.get()
        if name and type_ and source:
            self.result = {'name': name, 'type': type_, 'source': source}
            self.destroy()

    def get_column(self):
        return self.result
